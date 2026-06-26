"""
Airbnb Competitor Price Scraper
Ozone & Cassia Residences — Laguna Phuket

Usage:
    python3.12 airbnb_scraper.py                                     # all properties, default dates
    python3.12 airbnb_scraper.py --property ozone                    # only Ozone + its competitors
    python3.12 airbnb_scraper.py --checkin 2026-08-01 --checkout 2026-08-08 --property ozone
    python3.12 airbnb_scraper.py --dates-file dates.txt --workers 3
    python3.12 airbnb_scraper.py --visible                           # visible browser for debugging
"""

import asyncio, random, re, argparse, logging, json
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

from playwright.async_api import async_playwright, Page, TimeoutError as PWTimeout

try:
    from playwright_stealth import Stealth

    _stealth = Stealth()
    STEALTH_OK = True
except ImportError:
    STEALTH_OK = False
    print("⚠  playwright-stealth не установлен: pip install playwright-stealth")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font as XFont, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False
    print("pip install openpyxl")

# ── Logging ───────────────────────────────────────────────────────────────────
_HERE = Path(__file__).parent
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(_HERE / "airbnb_scraper.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

# ═════════════════════════════════════════════════════════════════════════════
#  CONFIG
# ═════════════════════════════════════════════════════════════════════════════
CURRENCY = "THB"
DELAY_MIN = 2.0
DELAY_MAX = 4.0
PAGE_LOAD_WAIT = 3.5

PROPERTIES_FILE = Path(__file__).parent / "airbnb_properties.json"


def load_properties() -> dict:
    if not PROPERTIES_FILE.exists():
        log.error(f"Файл не найден: {PROPERTIES_FILE}")
        raise SystemExit(1)
    with open(PROPERTIES_FILE, encoding="utf-8") as f:
        return json.load(f)


PROPERTIES = load_properties()

USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
]


# ═════════════════════════════════════════════════════════════════════════════
#  DATA MODELS
# ═════════════════════════════════════════════════════════════════════════════
@dataclass
class ListingOffer:
    price_night: Optional[float] = None  # цена/ночь (base rate)
    price_total: Optional[float] = (
        None  # итого за весь период (nights + cleaning + service)
    )
    cleaning_fee: Optional[float] = None  # уборка отдельно
    service_fee: Optional[float] = None  # сбор Airbnb
    discount_pct: Optional[float] = None  # недельная/месячная скидка %
    cancellation: str = ""  # "Free cancellation before DD Mon" / "Flexible" / "Strict"
    refundable: Optional[bool] = (
        None  # True=Flexible/Moderate, False=Strict, None=unknown
    )
    currency: str = "THB"


@dataclass
class ListingResult:
    name: str = ""
    url: str = ""
    label: str = ""
    rating: Optional[float] = None
    reviews: int = 0
    checkin: str = ""
    checkout: str = ""
    adults: int = 2
    offer: Optional[ListingOffer] = None  # одна цена на листинг
    is_own: bool = False
    error: str = (
        ""  # "" | "not_available" | "no_prices" | "timeout" | "context_destroyed"
    )
    scrape_ts: str = field(
        default_factory=lambda: datetime.now().isoformat(timespec="seconds")
    )

    @property
    def display_name(self) -> str:
        return self.label or self.name


# ═════════════════════════════════════════════════════════════════════════════
#  BROWSER HELPERS
# ═════════════════════════════════════════════════════════════════════════════
async def new_context(playwright, headless: bool = True):
    browser = await playwright.chromium.launch(
        headless=headless,
        args=[
            "--disable-blink-features=AutomationControlled",
            "--no-sandbox",
            "--disable-gpu",
            "--window-size=1440,900",
        ],
    )
    ctx = await browser.new_context(
        user_agent=random.choice(USER_AGENTS),
        viewport={"width": 1440, "height": 900},
        locale="en-GB",
        timezone_id="Asia/Bangkok",
        extra_http_headers={
            "Accept-Language": "en-GB,en;q=0.9",
            "Referer": "https://www.airbnb.com/",
            "sec-fetch-site": "same-origin",
            "sec-fetch-mode": "navigate",
        },
    )
    return ctx


async def apply_stealth(page: Page):
    """Apply stealth patches — playwright-stealth if available, manual patches as fallback."""
    if STEALTH_OK:
        await _stealth.apply_stealth_async(page)
    else:
        await page.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {get: () => false});
            Object.defineProperty(navigator, 'plugins', {get: () => [1,2,3,4,5]});
            Object.defineProperty(navigator, 'languages', {get: () => ['en-GB','en']});
            window.chrome = {runtime: {}};
            Object.defineProperty(screen, 'colorDepth', {get: () => 24});
        """)


async def dismiss_popups(page: Page):
    """Закрыть попапы Airbnb: cookies, translation banner, sign-in, любые диалоги."""
    selectors = [
        # Cookie consent
        '[data-testid="accept-btn"]',
        'button:has-text("Accept all cookies")',
        'button:has-text("Accept all")',
        # Translation announce modal ("Translation on" / "This symbol shows...")
        '[data-testid="translation-announce-modal"] button',
        'button:has-text("Got it")',
        # Generic close buttons on dialogs
        'button[aria-label="Close"]',
        '[aria-label="Close"]',
        # Sign-in / promo dismissal
        'button:has-text("Continue")',
        'button:has-text("Not now")',
        'button:has-text("Skip")',
        '[aria-label="Dismiss"]',
    ]
    for sel in selectors:
        try:
            el = page.locator(sel).first
            if await el.is_visible(timeout=500):
                await el.click()
                await asyncio.sleep(0.3)
        except Exception:
            pass
    # Press Escape to close any remaining overlay (translation modal, etc.)
    try:
        await page.keyboard.press("Escape")
        await asyncio.sleep(0.2)
    except Exception:
        pass


async def human_delay():
    await asyncio.sleep(random.uniform(DELAY_MIN, DELAY_MAX))


# ═════════════════════════════════════════════════════════════════════════════
#  PRICE HELPERS
# ═════════════════════════════════════════════════════════════════════════════
def parse_price(text: str) -> Optional[float]:
    if not text:
        return None
    cleaned = re.sub(r"[^\d.]", "", text.replace(",", ""))
    try:
        v = float(cleaned)
        return v if v > 100 else None
    except ValueError:
        return None


def classify_cancel(text: str) -> Optional[bool]:
    """True=refundable, False=non-refundable, None=unknown."""
    if not text:
        return None
    low = text.lower()
    if any(
        k in low
        for k in ("free cancellation", "flexible", "moderate", "fully refundable")
    ):
        return True
    if any(
        k in low
        for k in (
            "strict",
            "non-refundable",
            "no refund",
            "not refundable",
            "super strict",
        )
    ):
        return False
    return None


# ═════════════════════════════════════════════════════════════════════════════
#  PAGE DATA EXTRACTION — text-based parsing after "Show price breakdown" click
# ═════════════════════════════════════════════════════════════════════════════
PAGE_DATA_JS = r"""
() => {
    // ── Sidebar ───────────────────────────────────────────────────────────
    var sidebar = document.querySelector(
        '[data-section-id="BOOK_IT_SIDEBAR"], [data-plugin-in-point-id="BOOK_IT_SIDEBAR"]'
    );
    var sidebarText = sidebar ? (sidebar.innerText || '').trim() : '';
    // Strip "price below average" banners that repeat and push real price out of 1200-char window
    sidebarText = sidebarText.replace(/Your price is below the \d+-day average\n?/gi, '');
    sidebarText = sidebarText.replace(/Цена ниже средней за \d+ дн[^\n]*\n?/gi, '');
    sidebarText = sidebarText.trim();

    // ── Breakdown modal — find price dialog (works in any locale) ──────────
    var breakdownText = '';
    var breakdownMarkers = [
        'nights x', 'Price details',
        'ночей x', 'Разбивка цены', 'Цена с учетом',  // Russian
        'noches x', 'Desglose de precio',               // Spanish
        'nuits x', 'Détail du prix',                    // French
    ];
    document.querySelectorAll('[role="dialog"]').forEach(function(d) {
        var t = (d.innerText || '').trim();
        var isPrice = breakdownMarkers.some(function(m) { return t.indexOf(m) !== -1; });
        if (isPrice) breakdownText = t;
    });

    // ── Title ─────────────────────────────────────────────────────────────
    var h1 = document.querySelector('h1');
    var title = h1 ? (h1.innerText || '').trim() : document.title.split('|')[0].trim();

    // ── Rating / reviews (aria-label on review badge) ─────────────────────
    var ratingText = '';
    document.querySelectorAll('[aria-label]').forEach(function(el) {
        var a = el.getAttribute('aria-label') || '';
        if (/\d+(\.\d+)?\s*(out of 5|stars?)/i.test(a)) ratingText = a;
    });

    // ── Availability + special states ─────────────────────────────────────
    var bodyText = (document.body.innerText || '').toLowerCase();
    var unavailPhrases = [
        "these dates aren't available",
        "those dates are not available",
        "unavailable for your selected dates",
        "not available for your dates",
        "this place isn't available",
        "this place is not available",
        "sold out",
        "no longer available",
        "эти даты недоступны",
        "эти даты недоступн",
        "жильё недоступно",
        "это жильё недоступно",
        "недоступно для выбранных дат",
        "ไม่ว่างในช่วงเวลาที่คุณเลือก",
    ];
    var unavailable = unavailPhrases.some(function(p) { return bodyText.indexOf(p) !== -1; });

    // Min-stay violation: "X night minimum"
    var minStayMatch = bodyText.match(/(\d+)[- ]night minimum/);
    var minStay = minStayMatch ? parseInt(minStayMatch[1]) : 0;
    if (!minStay) {
        // Russian: "минимальный срок X ночей"
        var minStayRu = bodyText.match(/минимальный срок[^.]*?(\d+)\s*ноч/);
        if (minStayRu) minStay = parseInt(minStayRu[1]);
    }

    // Request-to-book (no instant pricing shown)
    var hasReserve = !!document.querySelector('[data-testid="book-it-cta"]');
    var hasRequest = !!document.querySelector('[data-testid="request-to-book-cta"]');
    var requestOnly = hasRequest && !hasReserve;

    return {
        sidebar:      sidebarText.substring(0, 1200),
        breakdown:    breakdownText.substring(0, 1200),
        title:        title.substring(0, 150),
        rating_text:  ratingText.substring(0, 80),
        unavailable:  unavailable,
        min_stay:     minStay,
        request_only: requestOnly,
    };
}
"""


def _extract_amount(text: str) -> Optional[float]:
    """Extract first THB amount from text, handles:
    - EU format: "1 200,68 ฿"  (space-thousands, comma-decimal, suffix ฿)
    - US format: "฿1,200.68"   (prefix ฿, comma-thousands, dot-decimal)
    - Integer:   "8 175 ฿"
    Returns negative value if text starts with minus sign.
    """
    if not text:
        return None
    t = text.strip()
    neg = bool(re.match(r"^[-−]", t))

    # EU decimal: "1 200,68 ฿" or "8 404,77 ฿"
    m = re.search(r"([\d][\d ]*\d|\d)\s*,\s*(\d{1,2})\s*฿", t)
    if m:
        v = float(m.group(1).replace(" ", "") + "." + m.group(2))
        return -v if neg else v

    # US prefix: "฿1,200.68"
    m = re.search(r"฿\s*([\d,]+(?:\.\d+)?)", t)
    if m:
        v = float(m.group(1).replace(",", ""))
        if v > 0:
            return -v if neg else v

    # Integer with ฿: "8 175 ฿" or "8175 ฿"
    m = re.search(r"([\d][\d ]{2,}\d|\d{4,})\s*฿", t)
    if m:
        v = float(m.group(1).replace(" ", ""))
        if v > 100:
            return -v if neg else v

    return None


def _all_amounts(text: str) -> list:
    """Return list of (position, value) for all ฿ prices found in text."""
    found = []
    seen_pos: set = set()

    def _add(pos: int, v: float) -> None:
        if v > 100 and not any(abs(pos - p) < 6 for p in seen_pos):
            found.append((pos, v))
            seen_pos.add(pos)

    for m in re.finditer(r"([\d][\d ]*\d|\d)\s*,\s*(\d{1,2})\s*฿", text):
        _add(m.start(), float(m.group(1).replace(" ", "") + "." + m.group(2)))
    for m in re.finditer(r"฿\s*([\d,]+(?:\.\d+)?)", text):
        v = float(m.group(1).replace(",", ""))
        _add(m.start(), v)
    for m in re.finditer(r"([\d][\d ]{2,}\d|\d{4,})\s*฿", text):
        _add(m.start(), float(m.group(1).replace(" ", "")))

    return sorted(found, key=lambda x: x[0])


def _parse_thb(text: str) -> Optional[float]:
    """Backward-compat alias for _extract_amount."""
    return _extract_amount(text)


def _parse_page_data(sidebar: str, breakdown: str, nights: int) -> dict:
    """
    Parse prices, cancellation, discount from plain-text sidebar + breakdown.
    Handles both EN (฿1,200.68) and RU/EU (1 200,68 ฿) number formats.
    """
    result: dict = {
        "price_night": None,
        "price_total": None,
        "cleaning_fee": None,
        "service_fee": None,
        "discount_pct": None,
        "cancellation": "",
    }

    # ── Total from sidebar ────────────────────────────────────────────────
    # Primary: look for ฿NUMBER near "total" keyword (works for EN pages).
    sidebar_total: Optional[float] = None
    m_sid = re.search(r"฿\s*([\d,]+(?:\.\d+)?)\s*total", sidebar, re.IGNORECASE)
    if m_sid:
        sidebar_total = float(m_sid.group(1).replace(",", ""))
    else:
        # Fallback: find all ฿ amounts (handles EU/RU format too), pick last/largest.
        sidebar_amounts = _all_amounts(sidebar)
        if sidebar_amounts:
            total_pos = sidebar.lower().find("total")
            if total_pos >= 0:
                before = [(p, v) for p, v in sidebar_amounts if p < total_pos + 20]
                sidebar_total = (before[-1][1] if before else sidebar_amounts[-1][1])
            else:
                sidebar_total = max(v for _, v in sidebar_amounts)
    result["price_total"] = sidebar_total

    # ── Cancellation from sidebar (EN + RU) ───────────────────────────────
    m = re.search(
        r"(Free cancellation[^\n]*"
        r"|[A-Z][a-z]+ cancellation[^\n]*"
        r"|Non-refundable[^\n]*"
        r"|Бесплатная отмена[^\n]*"
        r"|Без возврата[^\n]*)",
        sidebar,
        re.IGNORECASE,
    )
    if m:
        result["cancellation"] = m.group(1).strip()[:120]

    if not breakdown:
        return result

    lines = [ln.strip() for ln in breakdown.splitlines() if ln.strip()]
    breakdown_total: Optional[float] = None
    discount_amount: Optional[float] = None
    prev_is_discount = False

    for line in lines:
        low = line.lower()

        # Nightly rate: "7 nights x ฿1,221" or "7 ночей x 1 200,68 ฿"
        m_night = re.search(
            r"(\d+)\s*(?:nights?|ночей|noches|nuits?)\s*[x×]\s*(.*)",
            line,
            re.IGNORECASE,
        )
        if m_night and not result["price_night"]:
            v = _extract_amount(m_night.group(2))
            if not v:
                # Amount may be embedded differently — grab all amounts from line
                amounts = _all_amounts(line)
                v = amounts[-1][1] if amounts else None
            if v:
                result["price_night"] = v
            prev_is_discount = False
            continue

        # Cleaning fee: "Cleaning fee" / "уборк"
        if "cleaning" in low or "уборк" in low:
            v = _extract_amount(line)
            if v:
                result["cleaning_fee"] = v
            prev_is_discount = False
            continue

        # Service fee
        if "service fee" in low or "airbnb service" in low or "сервисный" in low:
            v = _extract_amount(line)
            if v:
                result["service_fee"] = v
            prev_is_discount = False
            continue

        # "Цена с учетом скидки" / "Price with discount" — label for final discounted total
        # Must be checked BEFORE the general discount check to avoid misclassification
        _is_total_label = (
            "учетом" in low  # RU: "цена с учетом скидки"
            or ("price" in low and ("discount" in low or "скидк" in low))
            or low.startswith("total")
            or low.startswith("итого")
        )

        # Discount label line: "Weekly discount -฿700 (10%)" or "Скидка за неделю"
        # Exclude total-label lines that happen to contain "скидк" as a genitive suffix
        if not _is_total_label and (
            "discount" in low or "скидк" in low or "descuento" in low
        ):
            m_pct = re.search(r"(\d+)\s*%", line)
            if m_pct:
                result["discount_pct"] = float(m_pct.group(1))
            v = _extract_amount(line)
            if v:
                discount_amount = abs(v)
                prev_is_discount = False
            else:
                prev_is_discount = True  # amount is on next line
            continue

        # Line immediately after discount label (amount only)
        if prev_is_discount:
            v = _extract_amount(line)
            if v:
                discount_amount = abs(v)
            prev_is_discount = False
            # fall through — this line may also be a standalone total candidate

        # Skip label-only lines (no price)
        line_amounts = _all_amounts(line)
        if not line_amounts:
            continue

        # Standalone amount line (pure number + ฿) → candidate for running total
        rest = re.sub(r"[-−\d\s,฿.]+", "", line).strip()
        if not rest:
            breakdown_total = line_amounts[-1][1]

    # ── Reconcile totals ──────────────────────────────────────────────────
    if breakdown_total:
        # Auto-compute discount % if we didn't find one explicitly
        if not result["discount_pct"]:
            if discount_amount and breakdown_total > 0:
                base = breakdown_total + discount_amount
                result["discount_pct"] = round(discount_amount / base * 100, 1)
            elif result["price_night"] and nights > 0:
                base = result["price_night"] * nights
                fees = (result["cleaning_fee"] or 0) + (result["service_fee"] or 0)
                stay_cost = breakdown_total - fees
                if base > stay_cost > 0:
                    result["discount_pct"] = round((base - stay_cost) / base * 100, 1)

        # Sidebar showed larger amount (original) → compute discount from difference
        if (
            sidebar_total
            and sidebar_total > breakdown_total
            and not result["discount_pct"]
        ):
            result["discount_pct"] = round(
                (sidebar_total - breakdown_total) / sidebar_total * 100, 1
            )

        result["price_total"] = breakdown_total

    # Fallback: compute nightly from total
    if result["price_total"] and not result["price_night"] and nights > 0:
        fees = (result["cleaning_fee"] or 0) + (result["service_fee"] or 0)
        result["price_night"] = round((result["price_total"] - fees) / nights, 2)

    return result


# ═════════════════════════════════════════════════════════════════════════════
#  SCRAPE ONE LISTING
# ═════════════════════════════════════════════════════════════════════════════
async def scrape_listing(
    page: Page,
    url: str,
    checkin: str,
    checkout: str,
    adults: int = 2,
    is_own: bool = False,
    label: str = "",
) -> ListingResult:

    result = ListingResult(
        url=url,
        checkin=checkin,
        checkout=checkout,
        adults=adults,
        is_own=is_own,
        label=label,
    )

    # Extract listing ID for clean URL building
    m = re.search(r"/rooms/(\d+)", url)
    listing_id = m.group(1) if m else url.split("/rooms/")[-1].split("?")[0]
    full_url = (
        f"https://www.airbnb.com/rooms/{listing_id}"
        f"?check_in={checkin}&check_out={checkout}&adults={adults}&currency={CURRENCY}"
    )

    try:
        log.info(f"  → {label or listing_id}")
        await page.goto(full_url, wait_until="domcontentloaded", timeout=40000)
        await asyncio.sleep(PAGE_LOAD_WAIT)
        await dismiss_popups(page)

        # Scroll down to trigger lazy-loading of price widget
        for _ in range(3):
            await page.evaluate("window.scrollBy(0, 400)")
            await asyncio.sleep(0.6)
        await page.evaluate("window.scrollTo(0, 0)")
        await asyncio.sleep(0.5)

        # Wait for booking sidebar to appear
        sidebar_loaded = False
        for sel in [
            '[data-section-id="BOOK_IT_SIDEBAR"]',
            '[data-testid="book-it-default"]',
            '[data-plugin-in-point-id="BOOK_IT_SIDEBAR"]',
        ]:
            try:
                await page.wait_for_selector(sel, timeout=8000)
                sidebar_loaded = True
                break
            except PWTimeout:
                continue

        if not sidebar_loaded:
            log.warning(f"    ⚠ Sidebar не загрузился: {label or listing_id}")

        nights = (
            datetime.strptime(checkout, "%Y-%m-%d")
            - datetime.strptime(checkin, "%Y-%m-%d")
        ).days or 1

        # Click "Check availability" if shown (dates not yet applied)
        try:
            btn = page.locator(
                'button:has-text("Check availability"), [data-testid="check-availability-cta"]'
            ).first
            if await btn.is_visible(timeout=2000):
                log.info("    → Кликаем Check availability")
                await btn.click()
                await asyncio.sleep(2.5)
        except Exception:
            pass

        try:
            await page.wait_for_load_state("networkidle", timeout=8000)
        except PWTimeout:
            await asyncio.sleep(1.5)

        # Dismiss any lingering modals (translation banner etc.) before interacting
        await dismiss_popups(page)
        await asyncio.sleep(0.5)

        # Click "Show price breakdown" / "Показать разбивку цены" (localized)
        try:
            breakdown_btn = page.locator(
                'button:has-text("Show price breakdown"), '
                'button:has-text("Показать разбивку цены"), '
                'button:has-text("price breakdown")'
            ).first
            if await breakdown_btn.is_visible(timeout=3000):
                await breakdown_btn.click()
                await asyncio.sleep(2.0)
                log.debug("    → Clicked breakdown button")
        except Exception:
            pass

        # ── Extract page data ──────────────────────────────────────────────
        page_data = {}
        for _attempt in range(3):
            try:
                page_data = await page.evaluate(PAGE_DATA_JS)
                break
            except Exception as e:
                if "Execution context was destroyed" in str(e) and _attempt < 2:
                    log.warning(f"    ↻ context destroyed, retry…")
                    await page.goto(
                        full_url, wait_until="domcontentloaded", timeout=35000
                    )
                    await asyncio.sleep(PAGE_LOAD_WAIT + 1.0)
                    await dismiss_popups(page)
                    try:
                        breakdown_btn = page.locator(
                            'button:has-text("Show price breakdown"), '
                            'button:has-text("Показать разбивку цены"), '
                            'button:has-text("price breakdown")'
                        ).first
                        if await breakdown_btn.is_visible(timeout=2000):
                            await breakdown_btn.click()
                            await asyncio.sleep(1.5)
                    except Exception:
                        pass
                else:
                    raise

        # Availability / special-state checks (none of these benefit from retry)
        if page_data.get("unavailable"):
            log.info(f"    ℹ Недоступно: {label or listing_id}")
            result.error = "not_available"
            return result
        if page_data.get("min_stay"):
            min_n = page_data["min_stay"]
            log.info(f"    ℹ Мин. срок {min_n} ночей: {label or listing_id}")
            result.error = f"min_stay_{min_n}"
            return result
        if page_data.get("request_only"):
            log.info(f"    ℹ Только запрос: {label or listing_id}")
            result.error = "request_only"
            return result

        # Parse structured data from text
        parsed = _parse_page_data(
            sidebar=page_data.get("sidebar", ""),
            breakdown=page_data.get("breakdown", ""),
            nights=nights,
        )

        price_night = parsed["price_night"]
        price_total = parsed["price_total"]
        cleaning_fee = parsed["cleaning_fee"]
        service_fee = parsed["service_fee"]
        discount_pct = parsed["discount_pct"]
        cancellation = parsed["cancellation"]

        # Title from page
        title = page_data.get("title", "") or ""
        result.name = (
            title.split("·")[0].strip() or (await page.title()).split("|")[0].strip()
        )

        # Rating from aria-label: "4.8 out of 5 stars, 42 reviews"
        rating_text = page_data.get("rating_text", "")
        if rating_text:
            m_r = re.search(r"(\d+\.?\d*)\s*out of 5", rating_text)
            m_rev = re.search(r"(\d+)\s*review", rating_text)
            if m_r:
                v = float(m_r.group(1))
                if 1 <= v <= 5:
                    result.rating = v
            if m_rev:
                result.reviews = int(m_rev.group(1))

        if not price_night and not price_total:
            log.warning(f"    ⚠ Нет цен: {result.display_name or listing_id}")
            log.warning(f"      sidebar={page_data.get('sidebar','')[:300]!r}")
            log.warning(f"      breakdown={page_data.get('breakdown','')[:300]!r}")
            result.error = "no_prices"
            return result

        result.offer = ListingOffer(
            price_night=price_night,
            price_total=price_total,
            cleaning_fee=cleaning_fee,
            service_fee=service_fee,
            discount_pct=discount_pct,
            cancellation=cancellation[:120],
            refundable=classify_cancel(cancellation),
            currency=CURRENCY,
        )

        nights_count = (
            datetime.strptime(checkout, "%Y-%m-%d")
            - datetime.strptime(checkin, "%Y-%m-%d")
        ).days or 1
        _adr = round(price_total / nights_count) if price_total else None
        _parts = [f"цена={round(price_total):,}฿"] if price_total else []
        if _adr:
            _parts.append(f"adr={_adr:,}฿")
        if price_night and discount_pct:
            _parts.append(f"до скидки={round(price_night * nights_count):,}฿")
        if discount_pct:
            _parts.append(f"скидка={discount_pct:.0f}%")
        if cleaning_fee:
            _parts.append(f"fees={round(cleaning_fee):,}฿")
        log.info(f"    ✓ {result.display_name or result.name}: " + "  ".join(_parts))
        return result

    except PWTimeout:
        result.error = "timeout"
        log.error(f"    ✗ Timeout [{label or listing_id}]: {url}")
        return result
    except Exception as e:
        err_str = str(e)
        result.error = (
            "context_destroyed"
            if "Execution context was destroyed" in err_str
            else err_str[:120]
        )
        log.error(f"    ✗ Ошибка [{label or listing_id}]: {e}")
        return result


# ═════════════════════════════════════════════════════════════════════════════
#  PARALLEL RUNNER
# ═════════════════════════════════════════════════════════════════════════════
async def scrape_batch(
    pw, jobs: list[dict], headless: bool, sem: asyncio.Semaphore
) -> list:
    async def _worker(job):
        async with sem:
            ctx = await new_context(pw, headless=headless)
            page = await ctx.new_page()
            await apply_stealth(page)
            try:
                await asyncio.sleep(random.uniform(0.5, 2.0))
                return await scrape_listing(
                    page,
                    job["url"],
                    job["checkin"],
                    job["checkout"],
                    job.get("adults", 2),
                    job.get("is_own", False),
                    job.get("label", ""),
                )
            finally:
                await ctx.close()

    return list(await asyncio.gather(*[_worker(j) for j in jobs]))


# ═════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ═════════════════════════════════════════════════════════════════════════════
def export_excel(results: list, output_path: str):
    if not EXCEL_OK:
        _csv(results, output_path.replace(".xlsx", ".csv"))
        return

    def fill(c):
        return PatternFill("solid", fgColor=c)

    def fnt(bold=False, color="000000", size=10, italic=False, underline=None):
        return XFont(
            name="Arial",
            bold=bold,
            color=color,
            size=size,
            italic=italic,
            underline=underline,
        )

    def aln(h="center", wrap=False):
        return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

    _thin = Side(style="thin", color="BDD7EE")
    _thick = Side(style="medium", color="2E75B6")

    def bdr(top=False, bottom=False, left=False, right=False):
        return Border(
            top=_thick if top else _thin,
            bottom=_thick if bottom else _thin,
            left=_thick if left else _thin,
            right=_thick if right else _thin,
        )

    def apply_prop_border(ws, start_r, end_r, n_cols):
        for r in range(start_r, end_r + 1):
            for c in range(1, n_cols + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = bdr(
                    top=(r == start_r),
                    bottom=(r == end_r),
                    left=(c == 1),
                    right=(c == n_cols),
                )

    wb = Workbook()

    # ── Sheet 1: все тарифы ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Все тарифы"
    ws.sheet_view.showGridLines = False

    COLS = [
        "Объект",
        "Гостей",
        "Цена ฿",
        "ADR ฿",
        "До скидки ฿",
        "Скидка %",
        "Fees ฿",
        "Возвратность",
        "vs Мой мин %",
    ]
    N = len(COLS)

    by_date: dict[str, list] = {}
    for r in results:
        by_date.setdefault(f"{r.checkin} → {r.checkout}", []).append(r)

    for ci, h in enumerate(COLS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = fnt(bold=True, color="FFFFFF", size=9)
        c.fill = fill("2E75B6")
        c.alignment = aln(wrap=True)
        c.border = bdr()
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "B2"

    row = 2
    for date_key, date_results in by_date.items():
        ws.merge_cells(f"A{row}:{get_column_letter(N)}{row}")
        c = ws.cell(row=row, column=1, value=f"📅  {date_key}")
        c.font = fnt(bold=True, color="FFFFFF", size=11)
        c.fill = fill("1F3864")
        c.alignment = aln(h="left")
        ws.row_dimensions[row].height = 24
        row += 1

        own_props = [
            r for r in date_results if r.is_own and r.offer and r.offer.price_total
        ]
        own_min_total = min((r.offer.price_total for r in own_props), default=None)

        ci_str, co_str = date_key.split(" → ")
        nights = (
            datetime.strptime(co_str, "%Y-%m-%d")
            - datetime.strptime(ci_str, "%Y-%m-%d")
        ).days or 1

        sorted_r = sorted(
            date_results,
            key=lambda r: (
                0 if r.is_own else 1,
                (r.offer.price_total if r.offer else None) or 999999,
            ),
        )

        for prop in sorted_r:
            is_own = prop.is_own
            prop_bg = "D5E8D4" if is_own else "FFFFFF"
            prop_fg = "1E4620" if is_own else "000000"
            name_str = f"{'★ ' if is_own else ''}{prop.display_name or prop.name}"

            _rid = re.search(r"/rooms/(\d+)", prop.url)
            prop_link = (
                f"https://www.airbnb.com/rooms/{_rid.group(1)}"
                f"?check_in={prop.checkin}&check_out={prop.checkout}&adults={prop.adults}&currency={CURRENCY}"
                if _rid
                else prop.url
            )

            prop_start = row

            if not prop.offer:
                err_vals = [name_str, prop.adults, f"— {prop.error or 'нет данных'}"] + [""] * (N - 3)
                for ci2, val in enumerate(err_vals, 1):
                    c = ws.cell(row=row, column=ci2)
                    if ci2 == 1:
                        c.value = name_str
                        c.hyperlink = prop_link
                        c.font = fnt(italic=True, color="0563C1", size=9, underline="single")
                    else:
                        c.value = val
                        c.font = fnt(italic=True, color="AA6600", size=9)
                    c.fill = fill("FFF2CC")
                    c.alignment = aln(h="left" if ci2 == 1 else "center")
                ws.row_dimensions[row].height = 17
                row += 1
                apply_prop_border(ws, prop_start, row - 1, N)
                continue

            o = prop.offer

            # vs Мой мин % — по итоговой цене
            vs_str, vs_color = "", "000000"
            if o.price_total and not is_own and own_min_total:
                diff = (o.price_total - own_min_total) / own_min_total
                vs_str = f"{diff:+.1%}"
                vs_color = "1A7A34" if diff > 0 else "C0392B"

            if o.refundable is True:
                cancel_str = o.cancellation or "Free cancellation"
                cancel_color = "1A7A34"
            elif o.refundable is False:
                cancel_str = o.cancellation or "Strict"
                cancel_color = "C0392B"
            else:
                cancel_str = o.cancellation[:70] or "—"
                cancel_color = "555555"

            disc_str = f"-{o.discount_pct:.0f}%" if o.discount_pct else ""
            adr = round(o.price_total / nights) if o.price_total else None
            # "До скидки ฿" = base accommodation cost before weekly/monthly discount
            pre_discount = (
                round(o.price_night * nights) if (o.price_night and o.discount_pct) else None
            )

            row_vals = [
                name_str,       # 1 Объект
                prop.adults,    # 2 Гостей
                o.price_total,  # 3 Цена ฿
                adr,            # 4 ADR ฿
                pre_discount,   # 5 До скидки ฿
                disc_str,       # 6 Скидка %
                o.cleaning_fee, # 7 Fees ฿
                cancel_str,     # 8 Возвратность
                vs_str,         # 9 vs Мой мин %
            ]

            for ci2, val in enumerate(row_vals, 1):
                c = ws.cell(row=row, column=ci2, value=val)
                c.fill = fill(prop_bg)
                c.border = bdr()
                c.alignment = aln(h="left" if ci2 in (1, 8) else "center")

                if ci2 == 1:
                    c.value = name_str
                    c.hyperlink = prop_link
                    c.font = fnt(bold=is_own, color="0563C1", size=9, underline="single")
                elif ci2 == 2:
                    c.font = fnt(color=prop_fg, size=9)
                elif ci2 == 3 and val:
                    c.font = fnt(bold=True, color=prop_fg, size=9)
                    c.number_format = '#,##0 "฿"'
                elif ci2 == 4 and val:
                    c.font = fnt(color="2E4057", size=9)
                    c.number_format = '#,##0 "฿"'
                elif ci2 == 5 and val:
                    c.font = fnt(color="999999", size=9, italic=True)
                    c.number_format = '#,##0 "฿"'
                elif ci2 == 6:
                    c.font = fnt(bold=True, color="E74C3C", size=9)
                elif ci2 == 7 and val:
                    c.font = fnt(color="666666", size=9, italic=True)
                    c.number_format = '#,##0 "฿"'
                elif ci2 == 8:
                    c.font = fnt(color=cancel_color, size=9)
                elif ci2 == 9:
                    c.font = fnt(bold=True, color=vs_color, size=9)
                else:
                    c.font = fnt(color=prop_fg, size=9)

            ws.row_dimensions[row].height = 17
            row += 1
            apply_prop_border(ws, prop_start, row - 1, N)

        row += 2

    # Объект | Гостей | Цена ฿ | ADR ฿ | До скидки ฿ | Скидка % | Fees ฿ | Возвратность | vs Мой мин %
    widths = [30, 8, 12, 10, 13, 9, 10, 45, 12]
    for ci2, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci2)].width = w

    # ── Sheet 2: Сводка ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Сводка")
    ws2.sheet_view.showGridLines = False
    SUM_COLS = [
        "Дата",
        "Объект",
        "Цена ฿",
        "ADR ฿",
        "До скидки ฿",
        "Скидка %",
        "vs Мой мин %",
    ]
    for ci2, h in enumerate(SUM_COLS, 1):
        c = ws2.cell(row=1, column=ci2, value=h)
        c.font = fnt(bold=True, color="FFFFFF", size=9)
        c.fill = fill("2E75B6")
        c.border = bdr()
    ws2.row_dimensions[1].height = 24

    srow = 2
    for date_key, date_results in by_date.items():
        own_min_total = next(
            (
                r.offer.price_total
                for r in date_results
                if r.is_own and r.offer and r.offer.price_total
            ),
            None,
        )
        for prop in sorted(
            date_results,
            key=lambda r: (
                0 if r.is_own else 1,
                (r.offer.price_total if r.offer else None) or 999999,
            ),
        ):
            is_own = prop.is_own
            bg = "D5E8D4" if is_own else ("FFFFFF" if srow % 2 == 0 else "F0F4FA")
            vs_str, vs_color = "", "000000"
            o = prop.offer
            if o and o.price_total and not is_own and own_min_total:
                diff = (o.price_total - own_min_total) / own_min_total
                vs_str = f"{diff:+.1%}"
                vs_color = "1A7A34" if diff > 0 else "C0392B"
            s_nights = (
                datetime.strptime(prop.checkout, "%Y-%m-%d")
                - datetime.strptime(prop.checkin, "%Y-%m-%d")
            ).days or 1
            s_adr = round(o.price_total / s_nights) if (o and o.price_total) else None
            s_pre = (
                round(o.price_night * s_nights)
                if (o and o.price_night and o.discount_pct)
                else None
            )
            row_data = [
                prop.checkin,
                f"{'★ ' if is_own else ''}{prop.display_name or prop.name}",
                o.price_total if o else None,
                s_adr,
                s_pre,
                f"-{o.discount_pct:.0f}%" if (o and o.discount_pct) else "",
                vs_str,
            ]
            for ci2, val in enumerate(row_data, 1):
                c = ws2.cell(row=srow, column=ci2, value=val)
                c.fill = fill(bg)
                c.border = bdr()
                c.font = fnt(
                    bold=is_own, color=vs_color if ci2 == 7 else "000000", size=9
                )
                c.alignment = aln(h="left" if ci2 == 2 else "center")
                if ci2 in (3, 4, 5) and val:
                    c.number_format = '#,##0 "฿"'
            ws2.row_dimensions[srow].height = 17
            srow += 1
        srow += 1

    for ci2, w in enumerate([14, 32, 12, 10, 13, 9, 12], 1):
        ws2.column_dimensions[get_column_letter(ci2)].width = w

    wb.save(output_path)
    log.info(f"✓ Сохранено: {output_path}")


def _csv(results, path):
    import csv

    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(
            [
                "name",
                "checkin",
                "checkout",
                "price_night",
                "price_total",
                "cleaning_fee",
                "service_fee",
                "discount_pct",
                "cancellation",
                "refundable",
            ]
        )
        for r in results:
            o = r.offer
            w.writerow(
                [
                    r.name,
                    r.checkin,
                    r.checkout,
                    o.price_night if o else "",
                    o.price_total if o else "",
                    o.cleaning_fee if o else "",
                    o.service_fee if o else "",
                    o.discount_pct if o else "",
                    o.cancellation if o else "",
                    o.refundable if o else "",
                ]
            )


# ═════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═════════════════════════════════════════════════════════════════════════════
async def run(
    date_pairs,
    prop_keys: list[str],
    adults=2,
    output="airbnb_analysis.xlsx",
    headless=True,
    workers=3,
):
    def build_jobs(checkin, checkout):
        jobs = []
        for key in prop_keys:
            prop = PROPERTIES[key]
            # Support both own_urls (list) and legacy url (string)
            own_entries = prop.get("own_urls") or [
                {"label": prop["label"], "url": prop["url"]}
            ]
            for own in own_entries:
                jobs.append(
                    {
                        "url": own["url"],
                        "checkin": checkin,
                        "checkout": checkout,
                        "adults": adults,
                        "is_own": True,
                        "label": own.get("label", prop["label"]),
                    }
                )
            for comp in prop.get("competitors", []):
                jobs.append(
                    {
                        "url": comp["url"],
                        "checkin": checkin,
                        "checkout": checkout,
                        "adults": adults,
                        "is_own": False,
                        "label": comp.get("label", ""),
                    }
                )
        return jobs

    total_jobs = sum(len(build_jobs(ci, co)) for ci, co in date_pairs)
    log.info(f"Всего задач: {total_jobs} | Воркеров: {workers}")

    RETRY_ERRORS = {"no_prices", "timeout", "context_destroyed"}
    all_results = []
    all_jobs_by_period = []

    async with async_playwright() as pw:
        sem = asyncio.Semaphore(workers)

        # Pass 1
        for checkin, checkout in date_pairs:
            log.info(f"\n{'─'*60}")
            log.info(f"📅  Период: {checkin} → {checkout}")
            log.info(f"{'─'*60}")
            jobs = build_jobs(checkin, checkout)
            batch = await scrape_batch(pw, jobs, headless=headless, sem=sem)
            all_jobs_by_period.append((jobs, batch))

        # Pass 2: retry failures once (not "not_available" / min_stay / request_only)
        for attempt in range(1, 2):
            failed = [
                (i, j)
                for i, (jobs, batch) in enumerate(all_jobs_by_period)
                for j, r in zip(jobs, batch)
                if r.error in RETRY_ERRORS
            ]
            if not failed:
                break
            labels = ", ".join(j["label"] or j["url"] for _, j in failed)
            log.info(f"\n↻ Повтор: {len(failed)} объектов ({labels})")
            await asyncio.sleep(4.0 * attempt)
            retry_jobs = [j for _, j in failed]
            retried = await scrape_batch(pw, retry_jobs, headless=headless, sem=sem)
            retry_map = {(r.url, r.checkin): r for r in retried}
            for i, (jobs, batch) in enumerate(all_jobs_by_period):
                all_jobs_by_period[i] = (
                    jobs,
                    [retry_map.get((r.url, r.checkin), r) for r in batch],
                )

        for _, batch in all_jobs_by_period:
            all_results.extend(batch)

    export_excel(all_results, output)

    print("\n" + "=" * 85)
    print(
        f"{'Дата':<13} {'Объект':<32} {'Цена/ночь':>10}  {'Итого':>10}  {'Уборка':>8}"
    )
    print("=" * 85)
    for r in all_results:
        mark = " ★" if r.is_own else ""
        name = (r.display_name or r.name)[:32] + mark
        if not r.offer:
            print(f"{r.checkin:<13} {name:<34} {'—':>10}  {'—':>10}  {r.error}")
            continue
        o = r.offer
        night_s = f"{o.price_night:.0f}฿" if o.price_night else "—"
        total_s = f"{o.price_total:.0f}฿" if o.price_total else "—"
        clean_s = f"{o.cleaning_fee:.0f}฿" if o.cleaning_fee else "—"
        disc_s = f" -{o.discount_pct:.0f}%" if o.discount_pct else ""
        print(
            f"{r.checkin:<13} {name:<34} {night_s:>10}  {total_s:>10}  {clean_s:>8}{disc_s}"
        )
    print(f"\n✓ Файл: {output}")


def parse_date_input(s: str) -> str:
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s.strip(), fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    raise ValueError(
        f"Неверный формат даты: {s!r}. Используй YYYY-MM-DD или DD.MM.YYYY"
    )


def prompt_dates() -> list[tuple[str, str]]:
    default_dates = PROPERTIES.get("default_dates", [])
    pairs = []
    print("\n─── Ввод дат ───────────────────────────────────────")
    if default_dates:
        print("Даты по умолчанию из airbnb_properties.json:")
        for i, d in enumerate(default_dates, 1):
            print(f"  {i}. {d['checkin']} → {d['checkout']}")
        print("Нажми Enter чтобы использовать их, или введи свои даты.")
    print("Формат: YYYY-MM-DD или DD.MM.YYYY")
    print("─────────────────────────────────────────────────────")

    while True:
        raw_ci = input("Заезд (checkin): ").strip()
        if not raw_ci:
            if default_dates:
                return [(d["checkin"], d["checkout"]) for d in default_dates]
            print("Нужно ввести хотя бы одну дату.")
            continue
        try:
            ci = parse_date_input(raw_ci)
        except ValueError as e:
            print(e)
            continue
        raw_co = input("Выезд (checkout): ").strip()
        try:
            co = parse_date_input(raw_co)
        except ValueError as e:
            print(e)
            continue
        pairs.append((ci, co))
        more = input("Добавить ещё период? (Enter — нет): ").strip()
        if not more:
            break
    return pairs


def build_date_pairs(args) -> list[tuple[str, str]]:
    if args.dates_file:
        pairs = []
        for line in Path(args.dates_file).read_text().splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.split()
            if len(parts) == 2:
                pairs.append((parts[0], parts[1]))
            elif len(parts) == 1:
                d = datetime.strptime(parts[0], "%Y-%m-%d").date()
                pairs.append((str(d), str(d + timedelta(days=1))))
        return pairs
    if args.checkin and args.checkout:
        return [(args.checkin, args.checkout)]
    return prompt_dates()


def main():
    p = argparse.ArgumentParser(description="Airbnb scraper — Laguna Phuket")
    prop_keys = [k for k in PROPERTIES if not k.startswith("_")]
    p.add_argument(
        "--property",
        choices=[*prop_keys, "all"],
        default="all",
        help=f"Объект: {', '.join(prop_keys)} или all (default: all)",
    )
    p.add_argument("--checkin")
    p.add_argument("--checkout")
    p.add_argument("--adults", type=int, default=2)
    p.add_argument("--dates-file", help="Файл с датами (YYYY-MM-DD YYYY-MM-DD)")
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M")
    p.add_argument("--output", default=f"airbnb_analysis_{ts}.xlsx")
    p.add_argument(
        "--workers",
        type=int,
        default=3,
        help="Параллельных браузеров (макс 8, default 3)",
    )
    p.add_argument("--visible", action="store_true", help="Показать браузер")
    args = p.parse_args()

    if args.workers > 8:
        args.workers = 8
        print("⚠  Снижено до 8 воркеров")

    selected_keys = (
        [k for k in prop_keys if k != "default_dates"]
        if args.property == "all"
        else [args.property]
    )
    pairs = build_date_pairs(args)

    if "airbnb_analysis_" in args.output:
        prop_part = args.property if args.property != "all" else "all"
        out_dir = _HERE / "output"
        out_dir.mkdir(exist_ok=True)
        args.output = str(out_dir / f"airbnb_{prop_part}_{ts}.xlsx")

    print(f"Объекты: {[PROPERTIES[k]['label'] for k in selected_keys]}")
    print(f"Дат: {len(pairs)}  Воркеров: {args.workers}")
    for ci, co in pairs:
        print(f"  {ci} → {co}")

    asyncio.run(
        run(
            date_pairs=pairs,
            prop_keys=selected_keys,
            adults=args.adults,
            output=args.output,
            headless=not args.visible,
            workers=args.workers,
        )
    )


if __name__ == "__main__":
    main()
