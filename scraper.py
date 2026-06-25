"""
Booking.com Competitor Price Scraper
Ozone & Cassia Residences — Laguna Phuket
v4: manual competitor lists per property · discounted price · parallel workers

Usage:
    python3.12 scraper.py                                    # all properties, default dates
    python3.12 scraper.py --property ozone                   # only Ozone + its competitors
    python3.12 scraper.py --property cassia                  # only Cassia + its competitors
    python3.12 scraper.py --checkin 2026-08-05 --checkout 2026-08-16 --property ozone
    python3.12 scraper.py --dates-file dates.txt --workers 2
    python3.12 scraper.py --dates-file dates.txt
    python3.12 scraper.py              # без дат — попросит ввести
"""

import asyncio, random, re, argparse, logging, json
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

from playwright.async_api import async_playwright, Page, TimeoutError as PWTimeout

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font as XFont, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False
    print("pip install openpyxl")

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("scraper.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

# ═════════════════════════════════════════════════════════════════════════════
#  CONFIG — РЕДАКТИРУЙ ЗДЕСЬ
# ═════════════════════════════════════════════════════════════════════════════

CURRENCY = "THB"
DELAY_MIN = 1.5
DELAY_MAX = 3.0
PAGE_LOAD_WAIT = 2.5

PROPERTIES_FILE = Path(__file__).parent / "properties.json"


def load_properties() -> dict:
    if not PROPERTIES_FILE.exists():
        log.error(f"Файл не найден: {PROPERTIES_FILE}")
        raise SystemExit(1)
    with open(PROPERTIES_FILE, encoding="utf-8") as f:
        return json.load(f)


PROPERTIES = load_properties()

USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
]


# ═════════════════════════════════════════════════════════════════════════════
#  DATA MODELS
# ═════════════════════════════════════════════════════════════════════════════
@dataclass
class RoomOffer:
    room_type: str = ""
    price_original: Optional[float] = None  # зачёркнутая цена (до скидки)
    price_final: Optional[float] = None  # цена со скидкой (или обычная если скидки нет)
    discount_pct: Optional[float] = None  # % скидки если есть
    cancellation: str = ""
    refundable: Optional[bool] = None
    breakfast: bool = False
    guests: Optional[int] = None
    size_m2: Optional[int] = None
    currency: str = "THB"

    @property
    def price_night(self) -> Optional[float]:
        """Актуальная цена — со скидкой если есть, иначе обычная."""
        return self.price_final or self.price_original


@dataclass
class PropertyResult:
    name: str = ""
    url: str = ""
    label: str = ""
    rating: Optional[float] = None
    reviews: int = 0
    checkin: str = ""
    checkout: str = ""
    adults: int = 2
    offers: list = field(default_factory=list)
    is_own: bool = False
    error: str = ""
    scrape_ts: str = field(
        default_factory=lambda: datetime.now().isoformat(timespec="seconds")
    )

    @property
    def min_price(self) -> Optional[float]:
        p = [o.price_night for o in self.offers if o.price_night]
        return min(p) if p else None

    @property
    def max_price(self) -> Optional[float]:
        p = [o.price_night for o in self.offers if o.price_night]
        return max(p) if p else None

    @property
    def display_name(self) -> str:
        return self.label or self.name


# ═════════════════════════════════════════════════════════════════════════════
#  BROWSER HELPERS
# ═════════════════════════════════════════════════════════════════════════════
async def new_context(playwright, headless: bool = True):
    browser = await playwright.chromium.launch(
        headless=headless,
        args=[
            "--disable-blink-features=AutomationControlled",
            "--no-sandbox",
            "--disable-gpu",
            "--window-size=1366,768",
        ],
    )
    ctx = await browser.new_context(
        user_agent=random.choice(USER_AGENTS),
        viewport={"width": 1366, "height": 768},
        locale="en-GB",
        timezone_id="Asia/Bangkok",
        extra_http_headers={"Accept-Language": "en-GB,en;q=0.9"},
    )
    await ctx.add_init_script(
        "Object.defineProperty(navigator,'webdriver',{get:()=>false});"
        "window.chrome={runtime:{}};"
    )
    return ctx


async def dismiss_all(page: Page):
    """Закрыть все попапы — cookie, sign-in, промо."""
    selectors = [
        "#onetrust-accept-btn-handler",
        '[data-testid="accept-button"]',
        'button[aria-label="Dismiss sign-in info."]',
        '[data-testid="dismissButton"]',
        ".modal-mask-closeBtn",
        'button[aria-label="Close"]',
        '[data-testid="modal-close-button"]',
        'button:has-text("Sign in later")',
        'button:has-text("Continue as guest")',
        'button:has-text("No thanks")',
    ]
    for sel in selectors:
        try:
            el = page.locator(sel).first
            if await el.is_visible(timeout=800):
                await el.click()
                await asyncio.sleep(0.4)
        except Exception:
            pass


async def human_delay():
    await asyncio.sleep(random.uniform(DELAY_MIN, DELAY_MAX))


# ═════════════════════════════════════════════════════════════════════════════
#  PRICE PARSER — учитывает зачёркнутую и финальную цены
# ═════════════════════════════════════════════════════════════════════════════
def parse_price(text: str) -> Optional[float]:
    if not text:
        return None
    cleaned = re.sub(r"[^\d.]", "", text.replace(",", ""))
    try:
        v = float(cleaned)
        return v if v > 100 else None
    except ValueError:
        return None


def extract_prices(price_block_html: str) -> tuple[Optional[float], Optional[float]]:
    """
    Из HTML блока цены извлекает (original, final).
    original = зачёркнутая цена (до скидки), может быть None
    final    = цена со скидкой (или единственная цена если скидки нет)
    """
    # Убираем HTML-теги чтобы получить текст
    text = re.sub(r"<[^>]+>", " ", price_block_html)
    # Находим все числа ≥ 100
    nums = [
        float(re.sub(r"[^\d.]", "", m))
        for m in re.findall(r"[\d,]{3,}(?:\.\d+)?", text)
    ]
    nums = [n for n in nums if n >= 100]

    if len(nums) == 0:
        return None, None
    if len(nums) == 1:
        return None, nums[0]
    # Два числа: обычно большее = зачёркнутая (оригинал), меньшее = финальная
    return max(nums), min(nums)


# ═════════════════════════════════════════════════════════════════════════════
#  DATE SETTING
# ═════════════════════════════════════════════════════════════════════════════
async def set_dates(page: Page, checkin: str, checkout: str, adults: int) -> bool:
    log.info("    📅 Setting dates...")

    for sel in [
        '[data-testid="searchbox-dates-container"]',
        'button[data-testid*="date"]',
    ]:
        try:
            el = page.locator(sel).first
            if await el.is_visible(timeout=2000):
                await el.click()
                await asyncio.sleep(2.0)
                break
        except Exception:
            pass

    ci = datetime.strptime(checkin, "%Y-%m-%d")
    co = datetime.strptime(checkout, "%Y-%m-%d")

    async def pick_day(dt: datetime) -> bool:
        ok = await page.evaluate(
            """(mv) => {
                var p = document.querySelector('#calendar_popup');
                if (p) p.style.display = 'block';
                var s = document.querySelector('#calendar_popup select, .bui-calendar select');
                if (!s) return false;
                s.value = mv;
                s.dispatchEvent(new Event('change', {bubbles:true}));
                return true;
            }""",
            f"{dt.year}-{dt.month}",
        )
        if not ok:
            return False
        await asyncio.sleep(2.0)

        return await page.evaluate(
            """([ds, dn]) => {
                var popup = document.querySelector('#calendar_popup');
                if (!popup) return false;
                var c = popup.querySelector('td[data-date="' + ds + '"]');
                if (!c) {
                    var tds = popup.querySelectorAll('td');
                    for (var i=0; i<tds.length; i++) {
                        if ((tds[i].getAttribute('onclick')||'').indexOf(ds) !== -1) { c=tds[i]; break; }
                    }
                }
                if (!c) {
                    var tds = popup.querySelectorAll('td');
                    for (var i=0; i<tds.length; i++) {
                        var t = (tds[i].innerText || tds[i].textContent || '').trim();
                        if (t === String(dn) && tds[i].className.indexOf('blocked') === -1) { c=tds[i]; break; }
                    }
                }
                if (c) { c.click(); return true; }
                return false;
            }""",
            [dt.strftime("%Y-%m-%d"), dt.day],
        )

    ci_ok = await pick_day(ci)
    await asyncio.sleep(0.4)
    co_ok = await pick_day(co)

    if ci_ok and co_ok:
        for sel in [
            'form[action*="searchresults"] button[type="submit"]',
            '[data-testid="searchbox-submit-button"]',
            '#frm button[type="submit"]',
        ]:
            try:
                btn = page.locator(sel).first
                if await btn.is_visible(timeout=1500):
                    await btn.click()
                    await page.wait_for_load_state("domcontentloaded", timeout=20000)
                    await asyncio.sleep(PAGE_LOAD_WAIT)
                    await dismiss_all(page)
                    return True
            except Exception:
                pass

    # Fallback: URL reload
    log.info("    → Calendar failed — URL fallback")
    base = re.sub(r"\.[a-z]{2}(-[a-z]{2})?\.html", ".html", page.url.split("?")[0])
    url = (
        f"{base}?checkin={checkin}&checkout={checkout}"
        f"&group_adults={adults}&no_rooms=1&selected_currency={CURRENCY}"
    )
    await page.goto(url, wait_until="domcontentloaded", timeout=30000)
    await asyncio.sleep(PAGE_LOAD_WAIT)
    await dismiss_all(page)
    return True


# ═════════════════════════════════════════════════════════════════════════════
#  SCRAPE ONE PROPERTY — все тарифы, зачёркнутые цены, возвратность
# ═════════════════════════════════════════════════════════════════════════════
async def scrape_property(
    page: Page,
    url: str,
    checkin: str,
    checkout: str,
    adults: int = 2,
    is_own: bool = False,
    label: str = "",
) -> PropertyResult:

    result = PropertyResult(
        url=url,
        checkin=checkin,
        checkout=checkout,
        adults=adults,
        is_own=is_own,
        label=label,
    )
    base = re.sub(r"\.[a-z]{2}(-[a-z]{2})?\.html", ".html", url.split("?")[0])
    full_url = (
        f"{base}?checkin={checkin}&checkout={checkout}"
        f"&group_adults={adults}&no_rooms=1&selected_currency={CURRENCY}&sb_price_type=total"
    )

    try:
        log.info(f"  → {(label or url[34:70])}")
        await page.goto(full_url, wait_until="domcontentloaded", timeout=35000)
        await asyncio.sleep(PAGE_LOAD_WAIT)
        await dismiss_all(page)

        # Property name
        for sel in [
            'h2[class*="pp-header__name"]',
            '[data-testid="property-header"] h2',
            "#hp_hotel_name",
            ".hp__hotel-name",
        ]:
            try:
                el = page.locator(sel).first
                if await el.is_visible(timeout=1000):
                    result.name = (await el.inner_text()).strip()
                    break
            except Exception:
                pass
        if not result.name:
            result.name = (await page.title()).split("–")[0].split("|")[0].strip()

        # Check if dates need to be set (only via visible button, no full page.content())
        needs_dates = False
        try:
            sp = page.locator(
                'button:has-text("Show prices"), [data-testid="show-prices-button"]'
            ).first
            needs_dates = await sp.is_visible(timeout=1500)
        except Exception:
            pass

        if needs_dates:
            log.info("    ℹ Dates not applied — using calendar")
            await set_dates(page, checkin, checkout, adults)

        # Scroll to load rooms
        for _ in range(4):
            await page.evaluate("window.scrollBy(0, 500)")
            await asyncio.sleep(0.5)

        # Rating
        for sel in [
            ".ac4a7896c7",
            ".bui-review-score__badge",
            '[data-testid="review-score-right-component"] span',
        ]:
            try:
                txt = (await page.locator(sel).first.inner_text(timeout=1000)).strip()
                v = float(txt.replace(",", "."))
                if 1 <= v <= 10:
                    result.rating = v
                    break
            except Exception:
                pass

        # Wait for room table
        table_found = False
        for sel in [
            '[data-testid="availability-table-wrapper"]',
            '[data-testid="hprt-table"]',
            "#hprt-table",
            ".hprt-table",
        ]:
            try:
                await page.wait_for_selector(sel, timeout=7000)
                table_found = True
                log.debug(f"    ✓ Table found")
                break
            except PWTimeout:
                continue

        if not table_found:
            log.warning(f"    ⚠ No room table: {result.name}")
            try:
                safe = re.sub(r"[^\w]", "_", result.name or "unk")[:20]
                # await page.screenshot(path=f"debug_{safe}.png")
            except Exception:
                pass
            result.error = "no_room_table"
            return result

        # Scroll to TOP of room table so all rows stay in viewport
        await page.evaluate("""
            var el = document.querySelector(
                '[data-testid="availability-table-wrapper"], [data-testid="hprt-table"], #hprt-table'
            );
            if (el) { window.scrollTo(0, el.getBoundingClientRect().top + window.scrollY - 80); }
        """)
        try:
            await page.wait_for_selector(
                '[data-testid="cancellation-policy"]', timeout=3000
            )
        except PWTimeout:
            pass
        await asyncio.sleep(0.8)

        # ── Extract ALL tariffs with discount-aware price parsing ─────────────
        # Wait for page to fully settle — prevents "Execution context was destroyed" on JS redirects
        try:
            await page.wait_for_load_state("networkidle", timeout=8000)
        except PWTimeout:
            await asyncio.sleep(1.0)  # graceful fallback if still busy

        offers_raw = []
        for _attempt in range(3):
            try:
                offers_raw = await page.evaluate(r"""
        () => {
            var results = [];

            function getPrice(row) {
                // Priority: discounted price element > struck-through old + new > single price
                var finalEl = row.querySelector(
                    '[data-testid="price-and-discounted-price"] [class*="discounted"], ' +
                    '.prco-valign-middle-helper .bui-price-display__value:last-child, ' +
                    '[class*="discount"] [class*="price"]:last-child, ' +
                    '.bui-price-display__value'
                );
                var origEl = row.querySelector(
                    '[class*="strikethrough"], [class*="crossed"], s, del, ' +
                    '[aria-label*="Original price"], [class*="original-price"], ' +
                    '.prco-text-nowrap-helper s, .bui-price-display__original'
                );

                // Get ALL price-like spans in the block
                var priceBlock = row.querySelector(
                    '[data-testid="price-and-discounted-price"], ' +
                    '.hprt-price-block, [class*="price-block"]'
                );

                var origText  = origEl  ? (origEl.innerText  || origEl.textContent  || '').trim() : '';
                var finalText = finalEl ? (finalEl.innerText || finalEl.textContent || '').trim() : '';

                // Collect all numeric texts from price block
                var allPrices = [];
                if (priceBlock) {
                    priceBlock.querySelectorAll('*').forEach(function(el) {
                        var t = (el.innerText || el.textContent || '').trim();
                        if (!t) return;
                        if (/^[\\d,\\.\\s฿THB]+$/.test(t) && t.replace(/[^0-9]/g,'').length >= 3) {
                            var n = parseFloat(t.replace(/[^0-9.]/g,'').replace(',',''));
                            var struck = el.tagName === 'S' || el.tagName === 'DEL';
                            if (!struck) {
                                var style = window.getComputedStyle(el);
                                struck = (style.textDecoration || style.textDecorationLine || '').indexOf('line-through') !== -1;
                            }
                            if (n >= 100) allPrices.push({text: t, val: n, isStruck: struck});
                        }
                    });
                }

                return {
                    orig_text:   origText,
                    final_text:  finalText,
                    all_prices:  allPrices,
                    block_html:  priceBlock ? priceBlock.innerHTML : '',
                };
            }

            function htmlToText(el) {
                if (!el) return '';
                // Strip all tags from innerHTML — captures text nodes after <strong> that innerText misses
                return el.innerHTML.replace(/<[^>]+>/g, ' ').replace(/\s+/g, ' ').trim();
            }

            function getCancel(row) {
                // Try policy-title first — its innerHTML has full text including date after <strong>
                var titleEl = row.querySelector('[data-testid="policy-title"]');
                if (titleEl) {
                    var t = htmlToText(titleEl);
                    if (t.length > 2) return t.substring(0, 120);
                }
                var sels = [
                    '[data-testid="cancellation-policy"]',
                    '[data-testid="cancellation-subtitle"]',
                    '.e2e-cancellation',
                    '.tpex-policy',
                    '.hprt-policy-name',
                    '.hprt-conditions li',
                ];
                for (var i = 0; i < sels.length; i++) {
                    var el = row.querySelector(sels[i]);
                    if (el) {
                        var t = htmlToText(el);
                        if (t.length > 2) return t.substring(0, 120);
                    }
                }
                return '';
            }

            function getGuests(row) {
                // Screen-reader span: "Number of guests: 2" or "Число гостей: 2"
                var srEl = row.querySelector('.bui-u-sr-only');
                if (srEl) {
                    var m = (srEl.textContent || '').match(/(\d+)/);
                    if (m) return parseInt(m[1]);
                }
                // Count occupancy icons (legacy table)
                var icons = row.querySelectorAll('.bicon-occupancy, .hprt-icon-adult, [data-testid="person-icon"]');
                if (icons.length > 0) return icons.length;
                // Modern table: occupancy element
                var occEl = row.querySelector('[data-testid="occupancy"], .c-occupancy-icons');
                if (occEl) {
                    var m2 = (occEl.textContent || '').match(/(\d+)/);
                    if (m2) return parseInt(m2[1]);
                }
                return null;
            }

            // Collect ALL cancellation policies from entire page in DOM order
            // (they may be in legacy #hprt-table, not inside availability-table-row)
            var pageCancels = [];
            document.querySelectorAll('[data-testid="cancellation-policy"]').forEach(function(el) {
                pageCancels.push(htmlToText(el).substring(0, 120));
            });

            function getSize(row) {
                // Look for m² in the entire row text
                var text = (row.innerText || row.textContent || '');
                var m = text.match(/(\d+)\s*m[²2]/i);
                if (m) return parseInt(m[1]);
                return null;
            }

            function classifyCancel(cancel) {
                var cl = cancel.toLowerCase();
                var isNR = cl.indexOf('non-refund') !== -1 || cl.indexOf('no refund') !== -1 ||
                           cl.indexOf('not refund') !== -1 || cl.indexOf('reschedule') !== -1;
                var isFR = !isNR && (cl.indexOf('free cancellation') !== -1 ||
                           cl.indexOf('free cancel') !== -1 || cl.indexOf('refundable') !== -1 ||
                           cl.indexOf('fully refund') !== -1);
                return isNR ? false : (isFR ? true : null);
            }

            // Modern rows
            var rows = document.querySelectorAll('[data-testid="availability-table-row"]');
            rows.forEach(function(row, idx) {
                var nameEl = row.querySelector(
                    '[data-testid="roomtype-name"], .hprt-roomtype-name, [class*="roomtype"]');
                var bfEl = row.querySelector('[data-testid="meal-plan"], [class*="meal"]');
                var p = getPrice(row);
                // Try cancel from row first, fallback to page-level collection by index
                var cancel = getCancel(row);
                if (!cancel && idx < pageCancels.length) cancel = pageCancels[idx];
                results.push({
                    room_type:  nameEl ? (nameEl.innerText || nameEl.textContent || '').trim() : '',
                    size_m2:    getSize(row),
                    cancel:     cancel,
                    refundable: classifyCancel(cancel),
                    breakfast:  bfEl ? (bfEl.innerText || '').toLowerCase().indexOf('breakfast') !== -1 : false,
                    guests:     getGuests(row),
                    orig_text:  p.orig_text,
                    final_text: p.final_text,
                    all_prices: p.all_prices,
                    block_html: p.block_html,
                });
            });

            // Legacy table fallback
            if (results.length === 0) {
                var legIdx = 0;
                document.querySelectorAll('#hprt-table tr, .hprt-table tr').forEach(function(row) {
                    var nameEl = row.querySelector('.hprt-roomtype-name, .jq_tooltip');
                    var p = getPrice(row);
                    if (!nameEl && !p.block_html) return;
                    var cancelEl = row.querySelector('[data-testid="cancellation-policy"], [data-testid="policy-title"], .e2e-cancellation, .hprt-conditions li');
                    var cancel = cancelEl ? htmlToText(cancelEl).substring(0, 120) : '';
                    if (!cancel && legIdx < pageCancels.length) cancel = pageCancels[legIdx];
                    results.push({
                        room_type:  nameEl ? (nameEl.innerText || nameEl.textContent || '').trim() : 'Room',
                        cancel:     cancel,
                        refundable: classifyCancel(cancel),
                        breakfast:  false,
                        size_m2:    getSize(row),
                        guests:     getGuests(row),
                        orig_text:  p.orig_text,
                        final_text: p.final_text,
                        all_prices: p.all_prices,
                        block_html: p.block_html,
                    });
                    legIdx++;
                });
            }

            return results;
        }
        """)
                break  # success
            except Exception as eval_err:
                if "Execution context was destroyed" in str(eval_err) and _attempt < 2:
                    log.warning(
                        f"    ↻ [{label or url[34:60]}] context destroyed, повторная навигация…"
                    )
                    try:
                        # Re-navigate — context is gone, waiting on broken page won't help
                        await page.goto(
                            full_url, wait_until="domcontentloaded", timeout=30000
                        )
                        await asyncio.sleep(PAGE_LOAD_WAIT + 1.5)
                        await dismiss_all(page)
                        try:
                            await page.wait_for_load_state("networkidle", timeout=6000)
                        except PWTimeout:
                            pass
                    except Exception:
                        await asyncio.sleep(2.0)
                else:
                    raise

        nights_count = (
            datetime.strptime(checkout, "%Y-%m-%d")
            - datetime.strptime(checkin, "%Y-%m-%d")
        ).days or 1
        min_total_price = (
            nights_count * 300
        )  # минимум 300฿/ночь — ниже явно ошибка парсинга

        seen = set()
        for raw in offers_raw:
            # Resolve prices: struck-through = original, lower = final
            all_p = raw.get("all_prices", [])
            struck = [x["val"] for x in all_p if x.get("isStruck")]
            unstruckt = [x["val"] for x in all_p if not x.get("isStruck")]

            orig_text = raw.get("orig_text", "")
            final_text = raw.get("final_text", "")

            price_orig = parse_price(orig_text) or (max(struck) if struck else None)
            price_final = parse_price(final_text) or (
                min(unstruckt) if unstruckt else None
            )

            # If only one price found in all_prices
            if not price_final and not price_orig and all_p:
                price_final = min(x["val"] for x in all_p)

            # Sanity: if final > orig, swap
            if price_orig and price_final and price_final > price_orig:
                price_orig, price_final = price_final, price_orig

            # If only one price exists, it's the final price
            if price_orig and not price_final:
                price_final = price_orig
                price_orig = None

            if not price_final:
                continue

            # Sanity check: price too low for a real stay — likely wrong DOM element
            if price_final < min_total_price:
                log.warning(
                    f"    ⚠ [{label}] подозрительно низкая цена {price_final}฿ "
                    f"за {nights_count} ночей ({raw.get('room_type','?')}), "
                    f"all_prices={[(x['val'], x['isStruck']) for x in all_p]}, "
                    f"block_html={raw.get('block_html','')[:120]}"
                )
                continue

            disc_pct = None
            if price_orig and price_final and price_orig > price_final:
                disc_pct = round((1 - price_final / price_orig) * 100)

            key = (raw.get("room_type", ""), price_final)
            if key in seen:
                continue
            seen.add(key)

            result.offers.append(
                RoomOffer(
                    room_type=raw.get("room_type", "") or "Room",
                    price_original=price_orig,
                    price_final=price_final,
                    discount_pct=disc_pct,
                    cancellation=raw.get("cancel", ""),
                    refundable=raw.get("refundable"),
                    breakfast=raw.get("breakfast", False),
                    guests=raw.get("guests"),
                    size_m2=raw.get("size_m2"),
                    currency=CURRENCY,
                )
            )

        if result.offers:
            disc_info = (
                f", {sum(1 for o in result.offers if o.discount_pct)} со скидкой"
            )
            log.info(
                f"    ✓ {result.display_name or result.name}: {len(result.offers)} тарифов, "
                f"мин={result.min_price}฿ макс={result.max_price}฿{disc_info}"
            )
        else:
            log.warning(f"    ⚠ Нет тарифов: {result.display_name or result.name}")
            for i, raw in enumerate(offers_raw):
                log.debug(
                    f"      raw[{i}]: room={raw.get('room_type','?')!r} "
                    f"orig={raw.get('orig_text','')!r} final={raw.get('final_text','')!r} "
                    f"all_prices={[(x['val'], x['isStruck']) for x in raw.get('all_prices',[])]} "
                    f"block_html={raw.get('block_html','')[:80]!r}"
                )
            result.error = "no_offers"

        return result

    except PWTimeout:
        result.error = "timeout"
        log.error(f"    ✗ Timeout [{label or url[34:60]}]: {url}")
        return result
    except Exception as e:
        err_str = str(e)
        result.error = "context_destroyed" if "Execution context was destroyed" in err_str else err_str[:120]
        log.error(f"    ✗ Ошибка [{label or url[34:60]}]: {e}")
        return result


# ═════════════════════════════════════════════════════════════════════════════
#  PARALLEL RUNNER
# ═════════════════════════════════════════════════════════════════════════════
async def scrape_batch(
    pw, jobs: list[dict], headless: bool, sem: asyncio.Semaphore
) -> list:
    async def _worker(job):
        async with sem:
            ctx = await new_context(pw, headless=headless)
            page = await ctx.new_page()
            try:
                await asyncio.sleep(random.uniform(0.3, 1.2))
                return await scrape_property(
                    page,
                    job["url"],
                    job["checkin"],
                    job["checkout"],
                    job.get("adults", 2),
                    job.get("is_own", False),
                    job.get("label", ""),
                )
            finally:
                await ctx.close()

    return list(await asyncio.gather(*[_worker(j) for j in jobs]))


# ═════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ═════════════════════════════════════════════════════════════════════════════
def export_excel(results: list, output_path: str):
    if not EXCEL_OK:
        _csv(results, output_path.replace(".xlsx", ".csv"))
        return

    def fill(c):
        return PatternFill("solid", fgColor=c)

    def fnt(bold=False, color="000000", size=10, italic=False, underline=None):
        return XFont(
            name="Arial",
            bold=bold,
            color=color,
            size=size,
            italic=italic,
            underline=underline,
        )

    def aln(h="center", wrap=False):
        return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

    _thin = Side(style="thin", color="BDD7EE")
    _thick = Side(style="medium", color="2E75B6")

    def bdr(top=False, bottom=False, left=False, right=False):
        """thin interior, thick on prop-group boundary sides."""
        return Border(
            top=_thick if top else _thin,
            bottom=_thick if bottom else _thin,
            left=_thick if left else _thin,
            right=_thick if right else _thin,
        )

    def apply_prop_border(ws, start_r, end_r, n_cols):
        """Re-apply borders so outer edges of property group are thick."""
        for r in range(start_r, end_r + 1):
            for c in range(1, n_cols + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = bdr(
                    top=(r == start_r),
                    bottom=(r == end_r),
                    left=(c == 1),
                    right=(c == n_cols),
                )

    wb = Workbook()

    # ── Sheet 1: все тарифы ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Все тарифы"
    ws.sheet_view.showGridLines = False

    COLS = [
        "Объект",
        "Тип номера / Тариф",
        "Гостей",
        "Цена ฿",
        "ADR ฿",
        "До скидки ฿",
        "Скидка %",
        "Возвратность",
        "vs Мой мин %",
    ]
    N = len(COLS)

    by_date: dict[str, list] = {}
    for r in results:
        by_date.setdefault(f"{r.checkin} → {r.checkout}", []).append(r)

    # Column headers — row 1, frozen so always visible while scrolling
    for ci, h in enumerate(COLS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = fnt(bold=True, color="FFFFFF", size=9)
        c.fill = fill("2E75B6")
        c.alignment = aln(wrap=True)
        c.border = bdr()
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "C2"  # freeze row 1 (column headers) — always visible

    row = 2
    for date_key, date_results in by_date.items():
        # Date header — bold separator, no repeated column headers
        ws.merge_cells(f"A{row}:{get_column_letter(N)}{row}")
        c = ws.cell(row=row, column=1, value=f"📅  {date_key}")
        c.font = fnt(bold=True, color="FFFFFF", size=11)
        c.fill = fill("1F3864")
        c.alignment = aln(h="left")
        ws.row_dimensions[row].height = 24
        row += 1

        own_props = [r for r in date_results if r.is_own]

        def _min_offers(refundable_flag):
            prices = [
                o.price_night
                for r in own_props
                for o in r.offers
                if o.price_night and o.refundable is refundable_flag
            ]
            return min(prices) if prices else None

        ref_refundable = _min_offers(True)  # мин возвратный своих
        ref_non_refundable = _min_offers(False)  # мин невозвратный своих
        ref = next(
            (r.min_price for r in date_results if r.is_own and r.min_price), None
        )
        sorted_r = sorted(
            date_results, key=lambda r: (0 if r.is_own else 1, r.min_price or 999999)
        )

        ci_str, co_str = date_key.split(" → ")
        nights = (
            datetime.strptime(co_str, "%Y-%m-%d")
            - datetime.strptime(ci_str, "%Y-%m-%d")
        ).days or 1

        for prop in sorted_r:
            is_own = prop.is_own
            prop_bg = "D5E8D4" if is_own else "FFFFFF"
            prop_fg = "1E4620" if is_own else "000000"
            name_str = f"{'★ ' if is_own else ''}{prop.display_name or prop.name}"

            # Build clickable URL with checkin/checkout dates
            base_url = re.sub(
                r"\.[a-z]{2}(-[a-z]{2})?\.html", ".html", prop.url.split("?")[0]
            )
            prop_link = (
                f"{base_url}?checkin={prop.checkin}&checkout={prop.checkout}"
                f"&group_adults={prop.adults}&no_rooms=1&selected_currency={CURRENCY}"
            )

            prop_start = row

            if not prop.offers:
                for ci, val in enumerate(
                    [
                        name_str,
                        f"— {prop.error or 'нет данных'}",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                    ],
                    1,
                ):
                    c = ws.cell(row=row, column=ci, value=val)
                    c.fill = fill("FFF2CC")
                    c.font = fnt(italic=True, color="AA6600", size=9)
                    c.alignment = aln(h="left" if ci <= 2 else "center")
                    if ci == 1:
                        c.hyperlink = prop_link
                        c.font = fnt(
                            italic=True, color="AA6600", size=9, underline="single"
                        )
                ws.row_dimensions[row].height = 17
                row += 1
                apply_prop_border(ws, prop_start, row - 1, N)
                continue

            offers_s = sorted(prop.offers, key=lambda o: o.price_night or 999999)
            for oi, offer in enumerate(offers_s):
                is_first = oi == 0
                bg = prop_bg if is_first else ("EAF4EA" if is_own else "F7FBFF")

                vs_str, vs_color = "", "000000"
                if offer.price_night and not is_own:
                    if offer.refundable is True:
                        cmp_ref = ref_refundable or ref
                    elif offer.refundable is False:
                        cmp_ref = ref_non_refundable or ref
                    else:
                        cmp_ref = ref
                    if cmp_ref:
                        diff = (offer.price_night - cmp_ref) / cmp_ref
                        vs_str = f"{diff:+.1%}"
                        vs_color = "1A7A34" if diff > 0 else "C0392B"

                if offer.refundable is True:
                    cancel_text = offer.cancellation or "Free cancellation"
                    ref_str = f"✓ {cancel_text[:55]}"
                    ref_color = "1A7A34"
                elif offer.refundable is False:
                    ref_str = "✗ Non-refundable"
                    ref_color = "C0392B"
                else:
                    ref_str, ref_color = offer.cancellation[:55] or "—", "555555"

                disc_str = f"-{offer.discount_pct:.0f}%" if offer.discount_pct else ""
                adr = round(offer.price_final / nights) if offer.price_final else None

                row_vals = [
                    name_str if is_first else "",  # 1
                    (
                        f"{offer.room_type} · {offer.size_m2}m²"
                        if offer.size_m2
                        else offer.room_type
                    )
                    or "—",  # 2
                    offer.guests,  # 3
                    offer.price_final,  # 4  Цена
                    adr,  # 5  ADR
                    (
                        offer.price_original
                        if offer.price_original != offer.price_final
                        else None
                    ),  # 6
                    disc_str,  # 7
                    ref_str,  # 8
                    vs_str,  # 9
                ]
                for ci, val in enumerate(row_vals, 1):
                    c = ws.cell(row=row, column=ci, value=val)
                    c.fill = fill(bg)
                    c.border = bdr()
                    c.alignment = aln(h="left" if ci in (1, 2, 8) else "center")
                    if ci == 1:
                        if is_first:
                            c.hyperlink = prop_link
                            c.font = fnt(
                                bold=is_own, color="0563C1", size=9, underline="single"
                            )
                        else:
                            c.font = fnt(color=prop_fg, size=9)
                    elif ci == 4:
                        c.font = fnt(bold=True, color=prop_fg, size=9)
                        c.number_format = '#,##0 "฿"'
                    elif ci == 5 and val:
                        c.font = fnt(bold=False, color="2E4057", size=9)
                        c.number_format = '#,##0 "฿"'
                    elif ci == 6 and val:
                        c.font = fnt(color="999999", size=9, italic=True)
                        c.number_format = '#,##0 "฿"'
                    elif ci == 7:
                        c.font = fnt(bold=True, color="E74C3C", size=9)
                    elif ci == 8:
                        c.font = fnt(color=ref_color, size=9)
                    elif ci == 9:
                        c.font = fnt(bold=True, color=vs_color, size=9)
                    else:
                        c.font = fnt(color=prop_fg, size=9)
                ws.row_dimensions[row].height = 17
                row += 1

            apply_prop_border(ws, prop_start, row - 1, N)

        row += 2

    widths = [28, 30, 7, 13, 11, 14, 9, 42, 12]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 2: Сводка ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Сводка")
    ws2.sheet_view.showGridLines = False
    SUM_COLS = [
        "Дата",
        "Объект",
        "Мин ฿",
        "Макс ฿",
        "Тарифов",
        "Со скидкой",
        "vs Мой мин %",
    ]
    for ci, h in enumerate(SUM_COLS, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = fnt(bold=True, color="FFFFFF", size=9)
        c.fill = fill("2E75B6")
        c.border = bdr()
    ws2.row_dimensions[1].height = 24

    srow = 2
    for date_key, date_results in by_date.items():
        ref = next(
            (r.min_price for r in date_results if r.is_own and r.min_price), None
        )
        for prop in sorted(
            date_results, key=lambda r: (0 if r.is_own else 1, r.min_price or 999999)
        ):
            is_own = prop.is_own
            bg = "D5E8D4" if is_own else ("FFFFFF" if srow % 2 == 0 else "F0F4FA")
            vs_str, vs_color = "", "000000"
            if ref and prop.min_price and not is_own:
                diff = (prop.min_price - ref) / ref
                vs_str = f"{diff:+.1%}"
                vs_color = "1A7A34" if diff > 0 else "C0392B"
            with_disc = sum(1 for o in prop.offers if o.discount_pct)
            row_data = [
                prop.checkin,
                f"{'★ ' if is_own else ''}{prop.display_name or prop.name}",
                prop.min_price,
                prop.max_price,
                len(prop.offers),
                with_disc or ("" if not with_disc else with_disc),
                vs_str,
            ]
            for ci, val in enumerate(row_data, 1):
                c = ws2.cell(row=srow, column=ci, value=val)
                c.fill = fill(bg)
                c.border = bdr()
                c.font = fnt(
                    bold=is_own, color=vs_color if ci == 7 else "000000", size=9
                )
                c.alignment = aln(h="left" if ci == 2 else "center")
                if ci in (3, 4) and val:
                    c.number_format = '#,##0 "฿"'
            ws2.row_dimensions[srow].height = 17
            srow += 1
        srow += 1

    for ci, w in enumerate([14, 30, 12, 12, 9, 10, 12], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    wb.save(output_path)
    log.info(f"✓ Сохранено: {output_path}")


def _csv(results, path):
    import csv

    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(
            [
                "name",
                "checkin",
                "checkout",
                "room_type",
                "price_final",
                "price_original",
                "discount_pct",
                "cancellation",
                "refundable",
                "breakfast",
            ]
        )
        for r in results:
            for o in r.offers:
                w.writerow(
                    [
                        r.name,
                        r.checkin,
                        r.checkout,
                        o.room_type,
                        o.price_final,
                        o.price_original,
                        o.discount_pct,
                        o.cancellation,
                        o.refundable,
                        o.breakfast,
                    ]
                )


# ═════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═════════════════════════════════════════════════════════════════════════════
async def run(
    date_pairs,
    prop_keys: list[str],
    adults=2,
    output="booking_analysis.xlsx",
    headless=True,
    workers=2,
):

    def build_jobs(checkin, checkout):
        jobs = []
        for key in prop_keys:
            prop = PROPERTIES[key]
            jobs.append(
                {
                    "url": prop["url"],
                    "checkin": checkin,
                    "checkout": checkout,
                    "adults": adults,
                    "is_own": True,
                    "label": prop["label"],
                }
            )
            for comp in prop["competitors"]:
                jobs.append(
                    {
                        "url": comp["url"],
                        "checkin": checkin,
                        "checkout": checkout,
                        "adults": adults,
                        "is_own": False,
                        "label": comp.get("label", ""),
                    }
                )
        return jobs

    total_jobs = sum(len(build_jobs(ci, co)) for ci, co in date_pairs)
    log.info(f"Всего задач: {total_jobs} | Воркеров: {workers}")

    MAX_RETRIES = 2
    RETRY_ERRORS = {"no_offers", "no_room_table", "timeout", "context_destroyed"}

    all_results = []
    all_jobs_by_period = []  # [(jobs, batch)] for retry pass

    async with async_playwright() as pw:
        sem = asyncio.Semaphore(workers)

        # ── Pass 1: scrape all periods without blocking on retries ─────────
        for checkin, checkout in date_pairs:
            log.info(f"\n{'─'*60}")
            log.info(f"📅  Период: {checkin} → {checkout}")
            log.info(f"{'─'*60}")
            jobs = build_jobs(checkin, checkout)
            batch = await scrape_batch(pw, jobs, headless=headless, sem=sem)
            all_jobs_by_period.append((jobs, batch))

        # ── Pass 2: retry all failures across all periods ──────────────────
        for attempt in range(1, MAX_RETRIES + 1):
            failed = [
                (i, j)
                for i, (jobs, batch) in enumerate(all_jobs_by_period)
                for j, r in zip(jobs, batch)
                if r.error in RETRY_ERRORS
            ]
            if not failed:
                break
            labels = ", ".join(j["label"] or j["url"][34:55] for _, j in failed)
            log.info(
                f"\n↻ Повтор {attempt}/{MAX_RETRIES}: {len(failed)} объектов ({labels})"
            )
            await asyncio.sleep(3.0 * attempt)
            retry_jobs = [j for _, j in failed]
            retried = await scrape_batch(pw, retry_jobs, headless=headless, sem=sem)
            retry_map = {(r.url, r.checkin): r for r in retried}
            # Patch results in-place
            for i, (jobs, batch) in enumerate(all_jobs_by_period):
                all_jobs_by_period[i] = (
                    jobs,
                    [retry_map.get((r.url, r.checkin), r) for r in batch],
                )

        for _, batch in all_jobs_by_period:
            all_results.extend(batch)

    export_excel(all_results, output)

    print("\n" + "=" * 85)
    print(f"{'Дата':<13} {'Объект':<30} {'Тарифов':>7}  {'Цена/ночь (финал)':>20}")
    print("=" * 85)
    for r in all_results:
        mark = " ★" if r.is_own else ""
        name = (r.display_name or r.name)[:30] + mark
        if not r.offers:
            print(f"{r.checkin:<13} {name:<32} {'—':>7}  нет данных")
            continue
        offers_s = sorted(r.offers, key=lambda o: o.price_night or 999999)
        print(f"{r.checkin:<13} {name:<32} {len(r.offers):>5}  мин={r.min_price:.0f}฿")
        for o in offers_s:
            if o.price_original and o.discount_pct:
                price_str = f"{o.price_final:.0f}฿  (было {o.price_original:.0f}฿, -{o.discount_pct:.0f}%)"
            else:
                price_str = f"{o.price_final:.0f}฿"
            if o.refundable is True:
                ref_s = f"✓ {o.cancellation or 'Free cancellation'}"
            elif o.refundable is False:
                ref_s = "✗ Non-refundable"
            else:
                ref_s = o.cancellation[:50] if o.cancellation else "—"
            guests_s = f"{o.guests}чел " if o.guests else ""
            print(
                f"  {'':<13} {'':<32}  └ {o.room_type[:35]:<36} {guests_s}{price_str:<30}  {ref_s}"
            )
    print(f"\n✓ Файл: {output}")


def parse_date_input(s: str) -> str:
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s.strip(), fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    raise ValueError(
        f"Неверный формат даты: {s!r}. Используй YYYY-MM-DD или DD.MM.YYYY"
    )


def prompt_dates() -> list[tuple[str, str]]:
    default_dates = PROPERTIES.get("default_dates", [])
    pairs = []
    print("\n─── Ввод дат ───────────────────────────────────────")
    if default_dates:
        print("Даты по умолчанию из properties.json:")
        for i, d in enumerate(default_dates, 1):
            print(f"  {i}. {d['checkin']} → {d['checkout']}")
        print("Нажми Enter чтобы использовать их, или введи свои даты.")
    print("Формат: YYYY-MM-DD или DD.MM.YYYY")
    print("Несколько периодов: вводи по одному, пустая строка — конец.")
    print("─────────────────────────────────────────────────────")

    while True:
        raw_ci = input("Заезд (checkin): ").strip()
        if not raw_ci:
            if default_dates:
                return [(d["checkin"], d["checkout"]) for d in default_dates]
            print("Нужно ввести хотя бы одну дату.")
            continue
        try:
            ci = parse_date_input(raw_ci)
        except ValueError as e:
            print(e)
            continue

        raw_co = input("Выезд (checkout): ").strip()
        try:
            co = parse_date_input(raw_co)
        except ValueError as e:
            print(e)
            continue

        pairs.append((ci, co))
        more = input("Добавить ещё период? (Enter — нет): ").strip()
        if not more:
            break

    return pairs


def build_date_pairs(args) -> list[tuple[str, str]]:
    if args.dates_file:
        pairs = []
        for line in Path(args.dates_file).read_text().splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.split()
            if len(parts) == 2:
                pairs.append((parts[0], parts[1]))
            elif len(parts) == 1:
                d = datetime.strptime(parts[0], "%Y-%m-%d").date()
                pairs.append((str(d), str(d + timedelta(days=1))))
        return pairs
    if args.checkin and args.checkout:
        return [(args.checkin, args.checkout)]
    # Нет дат в аргументах — спросить интерактивно
    return prompt_dates()


def main():
    p = argparse.ArgumentParser(description="Booking.com scraper — Laguna Phuket")
    p.add_argument(
        "--property",
        choices=[*PROPERTIES.keys(), "all"],
        default="all",
        help=f"Объект: {', '.join(PROPERTIES.keys())} или all (default: all)",
    )
    p.add_argument("--checkin")
    p.add_argument("--checkout")
    p.add_argument("--adults", type=int, default=2)
    p.add_argument("--dates-file", help="Файл с датами (YYYY-MM-DD YYYY-MM-DD)")
    p.add_argument("--output", default="booking_analysis.xlsx")
    p.add_argument(
        "--workers",
        type=int,
        default=4,
        help="Параллельных браузеров (макс 8, default 4)",
    )
    p.add_argument("--visible", action="store_true", help="Показать браузер")
    args = p.parse_args()

    if args.workers > 8:
        args.workers = 8
        print("⚠  Снижено до 8 воркеров")

    prop_keys = list(PROPERTIES.keys()) if args.property == "all" else [args.property]
    pairs = build_date_pairs(args)

    print(f"Объекты: {[PROPERTIES[k]['label'] for k in prop_keys]}")
    print(f"Дат: {len(pairs)}  Воркеров: {args.workers}")
    for ci, co in pairs:
        print(f"  {ci} → {co}")

    asyncio.run(
        run(
            date_pairs=pairs,
            prop_keys=prop_keys,
            adults=args.adults,
            output=args.output,
            headless=not args.visible,
            workers=args.workers,
        )
    )


if __name__ == "__main__":
    main()
